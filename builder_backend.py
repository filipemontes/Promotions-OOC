#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import io, re
from copy import copy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.table import Table, TableStyleInfo

# ========= Defaults & constants =========

DEFAULT_HOD_SPLIT_CANDIDATES = ["Manager Level 5 Name", "HOD/Manager"]
DEFAULT_L2_COL = "Manager Level 6 Name"   # Manager L+2

# Suggested pairs for auto-mapping (user can override in UI)
DEFAULT_MAPPING_PAIRS = [
    ("Employee ID", "Employee ID"),
    ("Preferred Name", "Employee Name"),
    ("Manager Level 6 Name", "Manager L+2"),
    ("Manager Level 7 Name", "Manager L+1"),
    ("Manager Level 5 Name", "HOD/Manager"),
    ("Division", "Team"),
    ("Service Date", "Start Date"),
    ("Last Base Pay Increase Date", "Last Base Pay Increase Date"),
    ("Location", "Location"),
    ("Job Title", "Current Role"),
    ("Job Title", "Role"),
    ("Compensation Grade", "Band"),
    ("Total Base Pay Annualized - Currency", "Currency"),
    ("Total Base Pay Amount (Local)", "Current Annual Salary"),
    ("Rating - Previous", "End-Year 2023"),
    ("Rating - Most Recent", "End-Year 2024"),
    ("Job Profile", "Job Profile"),
    ("Job profile", "Job Profile"),
    (" Job Family", "Job Family"),
    ("Job Family", "Job Family"),("Eligibility", "Eligibility"),("Last Promotion", "Last Promotion")
]

DEFAULT_NUMERIC_TARGETS = {"Current Annual Salary", "End-Year 2023", "End-Year 2024"}

HRV_NAME = "HR_View"              # used inside each per-HOD workbook
ALL_MGRS_TABLE = "All_Managers"   # table name in MasterFile
ALL_MGRS_SHEET = "All_Managers"   # sheet hosting All_Managers

# >>> FIXED columns for All_Managers sheet (as requested)
ALL_MANAGERS_FIXED_COLUMNS = [
    "Employee ID",
    "Manager Proposal",
    "Priority",
    "Pay change type",
    "Pay change reason",
    "Justification",
    "Role",
    "HOD/Manager",
    "Manager L+1",
    "Employee Name",
]

# ========= Helpers (pure) =========

def norm(s):
    if s is None: return ""
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s)
    return s.lower()

def build_src_lookup(columns):
    return {norm(c): c for c in columns}

def targets_by_candidates(mapping_pairs):
    mp = {}
    for src, tgt in mapping_pairs:
        mp.setdefault(tgt, []).append(src)
    return mp

def propose_mapping(target_headers, headcount_columns, mapping_pairs=DEFAULT_MAPPING_PAIRS):
    """
    Returns {target_header -> headcount_source_or_None} using:
    1) exact name match (case/space-insensitive)
    2) suggestion pairs
    """
    src_lookup = build_src_lookup(headcount_columns)
    mapping = {t: src_lookup.get(norm(t)) for t in target_headers}
    suggestions = targets_by_candidates(mapping_pairs)
    for tgt in target_headers:
        if mapping.get(tgt):
            continue
        for cand in suggestions.get(tgt, []):
            src = src_lookup.get(norm(cand))
            if src:
                mapping[tgt] = src
                break
    return mapping

def read_headcount(file_like):
    df = pd.read_excel(file_like, dtype=str)
    df.columns = [str(c) for c in df.columns]
    return df

def _row_has_any_values(ws, row_idx, max_cols=500):
    for c in range(1, min(ws.max_column, max_cols) + 1):
        if ws.cell(row_idx, c).value not in (None, ""):
            return True
    return False

def detect_data_sheet_and_header(
    wb,
    required=("Employee ID", "Employee Name"),
    exclude_titles=None,
):
    """
    Find the main data sheet by header row. Can skip auxiliary sheets like All_Managers / HR_View.

    Returns: (ws, header_row, subheader_row_or_None)
    """
    def _nt(x):
        return re.sub(r"\s+", " ", (x or "")).strip().lower()
    exclude_set = set(_nt(t) for t in (exclude_titles or []))

    want = set(norm(x) for x in required)
    best = None

    for ws in wb.worksheets:
        if _nt(ws.title) in exclude_set:
            continue
        max_rows = min(ws.max_row or 0, 60)
        for r in range(1, max_rows + 1):
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            present = set(norm(v) for v in vals if v is not None)
            score = len(want.intersection(present))
            if score == len(want):
                sub = r - 1 if r > 1 and _row_has_any_values(ws, r - 1) else None
                return ws, r, sub
            if best is None or score > best[2]:
                best = (ws, r, score)

    if best and best[2] > 0:
        ws, r, _ = best
        sub = r - 1 if r > 1 and _row_has_any_values(ws, r - 1) else None
        return ws, r, sub

    raise RuntimeError("Couldn't find the data header row (need 'Employee ID' & 'Employee Name').")

def header_positions(ws, header_row):
    pos = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is None: continue
        v = str(v).strip()
        pos[v] = (c, get_column_letter(c))
    return pos

def collect_model_formulas(ws, model_row):
    formulas = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(model_row, c).value
        if isinstance(val, str) and val.startswith("="):
            formulas[c] = val
    return formulas

def collect_model_number_formats(ws, model_row):
    fmts = {}
    for c in range(1, ws.max_column + 1):
        fmts[c] = ws.cell(model_row, c).number_format
    return fmts

def coerce_excel_number(v):
    if v is None: return v
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    if s == "": return s
    s2 = re.sub(r"[^\d,.\-]", "", s)
    if "." in s2 and "," in s2:
        try: return float(s2.replace(",", ""))                    # US
        except:
            try: return float(s2.replace(".", "").replace(",", "."))  # EU
            except: return v
    if "," in s2 and "." not in s2:
        try: return float(s2.replace(".", "").replace(",", "."))
        except: return v
    try:
        return float(s2)
    except:
        return v

def df_to_rows_via_mapping(df, target_headers, mapping, numeric_targets=None, date_targets=None):
    numeric_targets = set(numeric_targets or [])
    date_targets = set(date_targets or {"Start Date", "Last Base Pay Increase Date"})

    rows = []
    for _, rec in df.iterrows():
        out = {}
        for tgt in target_headers:
            src = mapping.get(tgt)
            if not src:
                continue
            v = rec.get(src, "")
            if pd.isna(v): v = ""
            if tgt in date_targets:
                parsed = pd.to_datetime(v, errors="coerce")
                if pd.notna(parsed):
                    v = parsed.to_pydatetime()
            if tgt in numeric_targets:
                v = coerce_excel_number(v)
            out[tgt] = v
        rows.append(out)
    return rows

def clear_cells_below(ws, header_row, tmpl_headers_pos, max_rows=100000):
    start_row = header_row + 1
    end_row   = min(start_row + max_rows, 1048576)
    cols = [col for (col, _) in tmpl_headers_pos.values()]
    for r in range(start_row, end_row):
        for col_idx in cols:
            ws.cell(r, col_idx).value = None

_cell_ref_re = re.compile(r"(?P<sheet>(?:'[^']+'|[A-Za-z0-9_\.]+)!)?(?P<col>\$?[A-Za-z]{1,3})(?P<row>\$?\d+)")

def _shift_formula_rows_regex(formula: str, delta: int) -> str:
    def repl(m):
        sheet, col, row = m.group("sheet"), m.group("col"), m.group("row")
        if row.startswith("$"):
            return (sheet or "") + col + row
        return (sheet or "") + col + str(int(row) + delta)
    return _cell_ref_re.sub(repl, formula)

def translate_row_formula(formula: str, origin_col_idx: int, origin_row: int, target_row: int) -> str:
    delta = target_row - origin_row
    try:
        origin_addr = f"{get_column_letter(origin_col_idx)}{origin_row}"
        return Translator(formula, origin=origin_addr).translate_row(delta)
    except Exception:
        return _shift_formula_rows_regex(formula, delta)

def clone_column_widths(src_ws, dst_ws):
    for col, dim in src_ws.column_dimensions.items():
        try:
            dst_ws.column_dimensions[col].width = dim.width
        except Exception:
            pass

def copy_row_values_and_styles(src_ws, dst_ws, src_row, dst_row):
    for c in range(1, src_ws.max_column + 1):
        sc = src_ws.cell(src_row, c)
        dc = dst_ws.cell(dst_row, c)
        dc.value = sc.value
        dc.font = copy(sc.font)
        dc.border = copy(sc.border)
        dc.fill = copy(sc.fill)
        dc.number_format = sc.number_format
        dc.protection = copy(sc.protection)
        dc.alignment = copy(sc.alignment)


def normalize_header_styles(
    ws,
    header_row,
    subheader_row=None,
    font_size=11,
    bold=True,
    center=True,
    row_height=15,
    header_font_color="FFFFFF",     # WHITE for dark header fill
    subheader_font_color="FFFFFF",  # WHITE for dark subheader fill
):
    """
    Sets clean, consistent formatting for header & optional sub-header rows:
    - white text (so it pops on dark fills),
    - font size ~11, bold on header,
    - centered header text, normal row heights,
    - preserves the cell fills/borders copied from the template.
    """
    # Header style
    hdr_font = Font(size=font_size, bold=bold, color=header_font_color)
    hdr_align = Alignment(
        horizontal="center" if center else None,
        vertical="center",
        wrap_text=True
    )
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, c)
        cell.font = hdr_font
        cell.alignment = hdr_align
    ws.row_dimensions[header_row].height = row_height

    # Sub-header style (if present)
    if subheader_row:
        sub_font = Font(size=font_size, bold=False, color=subheader_font_color)
        sub_align = Alignment(vertical="center", wrap_text=True)
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(subheader_row, c)
            cell.font = sub_font
            cell.alignment = sub_align
        ws.row_dimensions[subheader_row].height = row_height


def rename_headers_in_sheet(ws, header_row, rename_map):
    """rename_map: {original_header_text -> new_header_text}"""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is None: continue
        v = str(v).strip()
        if v in rename_map:
            ws.cell(header_row, c).value = rename_map[v]

def sanitize_sheet_name(name: str) -> str:
    s = re.sub(r"[:\\/?*\[\]]", "_", str(name))
    s = s.strip() or "Sheet"
    return s[:31]

def sanitize_filename(name: str) -> str:
    base = re.sub(r'[/\\:*?"<>|]', "_", str(name))
    base = base.strip() or "Unknown_HOD"
    return base[:200]

def remove_table_if_exists(ws, name: str):
    try:
        if hasattr(ws, "tables"):
            ws.tables.pop(name, None)
        else:
            if hasattr(ws, "_tables"):
                ws._tables = [t for t in ws._tables if getattr(t, "displayName", None) != name]
    except Exception:
        pass

def extract_hrview_columns_from_formulas(model_formulas: dict) -> list:
    cols = []
    seen = set()
    pat = re.compile(rf"{HRV_NAME}\[([^\]]+)\]")
    for f in model_formulas.values():
        if not isinstance(f, str): continue
        for col in pat.findall(f):
            key = col.strip()
            if key not in seen:
                cols.append(key)
                seen.add(key)
    if "Employee ID" in cols:
        cols.remove("Employee ID")
    return ["Employee ID"] + cols

def build_or_refresh_hr_view_table(wb, hr_cols: list, data_rows: list, numeric_targets=None):
    numeric_targets = set(numeric_targets or [])
    if HRV_NAME in wb.sheetnames:
        ws = wb[HRV_NAME]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=HRV_NAME)
    remove_table_if_exists(ws, HRV_NAME)

    if not hr_cols:
        hr_cols = ["Employee ID"]

    for j, colname in enumerate(hr_cols, start=1):
        ws.cell(1, j).value = colname

    for i, row in enumerate(data_rows, start=2):
        for j, colname in enumerate(hr_cols, start=1):
            val = row.get(colname, "")
            if isinstance(val, str) and colname in numeric_targets:
                val = coerce_excel_number(val)
            ws.cell(i, j).value = val

    last_col_letter = get_column_letter(len(hr_cols))
    last_row = 1 + len(data_rows)
    table_ref = f"A1:{last_col_letter}{max(last_row, 1)}"
    table = Table(displayName=HRV_NAME, ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

def write_rows(ws, tmpl_headers_pos, data_rows, model_formulas, model_row, model_numfmts, initialize_model_row=False):
    if initialize_model_row:
        for col_idx, f in model_formulas.items():
            ws.cell(model_row, col_idx).value = f
            try:
                ws.cell(model_row, col_idx).number_format = model_numfmts.get(col_idx, ws.cell(model_row, col_idx).number_format)
            except Exception:
                pass

    start_row = model_row
    for i, rowdict in enumerate(data_rows):
        r = start_row + i
        first = (i == 0)
        for hdr, (col_idx, _) in tmpl_headers_pos.items():
            target_cell = ws.cell(r, col_idx)
            if hdr in rowdict:
                target_cell.value = rowdict[hdr]
            else:
                if first:
                    pass
                else:
                    if col_idx in model_formulas:
                        base = model_formulas[col_idx]
                        translated = translate_row_formula(base, col_idx, model_row, r)
                        target_cell.value = translated
            try:
                target_cell.number_format = model_numfmts.get(col_idx, target_cell.number_format)
            except Exception:
                pass

def expand_table_if_present(ws, header_row, last_row, min_col=None, max_col=None):
    try:
        tables = list(getattr(ws, "tables", {}).values()) if hasattr(ws, "tables") else list(getattr(ws, "_tables", []))
    except Exception:
        tables = []
    if not tables:
        return
    for tbl in tables:
        ref = getattr(tbl, "ref", None)
        if not ref:
            continue
        m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", ref)
        if not m:
            continue
        c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
        if r1 != header_row:
            continue
        if min_col and max_col:
            c1 = get_column_letter(min_col)
            c2 = get_column_letter(max_col)
        new_ref = f"{c1}{header_row}:{c2}{max(last_row, header_row)}"
        try:
            tbl.ref = new_ref
        except Exception:
            pass

# ========= Template context =========

def load_template_context(uploaded_file, exclude_sheets=None):
    """
    Returns a dict with workbook context for a template uploaded in Streamlit.
    exclude_sheets: iterable of sheet names to ignore (e.g., {"All_Managers"} or {"HR_View"}).
    """
    keep_vba = False
    out_ext = ".xlsx"
    name = getattr(uploaded_file, "name", "template.xlsx").lower()
    if name.endswith(".xlsm"):
        keep_vba = True
        out_ext = ".xlsm"

    wb = load_workbook(io.BytesIO(uploaded_file.getvalue()), data_only=False, keep_vba=keep_vba)
    data_ws, header_row, subheader_row = detect_data_sheet_and_header(
        wb,
        required=("Employee ID", "Employee Name"),
        exclude_titles=set(exclude_sheets or []),
    )
    tmpl_pos = header_positions(data_ws, header_row)
    model_row = header_row + 1
    model_formulas = collect_model_formulas(data_ws, model_row)
    model_numfmts  = collect_model_number_formats(data_ws, model_row)

    return {
        "wb": wb,
        "keep_vba": keep_vba,
        "ext": out_ext,
        "data_ws": data_ws,
        "header_row": header_row,
        "subheader_row": subheader_row,
        "tmpl_pos": tmpl_pos,
        "model_row": model_row,
        "model_formulas": model_formulas,
        "model_numfmts": model_numfmts,
        "target_headers": list(tmpl_pos.keys()),
    }

# ========= Preview =========

def preview_dataframe(df, target_headers, mapping, numeric_targets):
    rows = df_to_rows_via_mapping(df, target_headers, mapping, numeric_targets=numeric_targets)
    if not rows:
        return pd.DataFrame(columns=target_headers)
    frame = pd.DataFrame(rows)
    for h in target_headers:
        if h not in frame.columns:
            frame[h] = ""
    return frame[target_headers]

# ========= Build functions =========

def build_per_hod_workbooks(
    df,
    template_file,
    mapping_perhod,
    hod_col,
    l2_col,
    numeric_targets,
    perhod_rename_map=None,
    normalize_headers=True,
    output_name_pattern="{HOD}",
):
    ctx = load_template_context(template_file)
    keep_vba = ctx["keep_vba"]
    out_ext  = ctx["ext"]

    src_headers = df.columns.tolist()
    if hod_col not in src_headers:
        raise RuntimeError(f"HOD split column '{hod_col}' not found in Headcount.")
    if l2_col not in src_headers:
        raise RuntimeError(f"L+2 column '{l2_col}' not found in Headcount.")

    unique_hods = sorted([x for x in df[hod_col].dropna().unique() if str(x).strip() != ""], key=lambda s: str(s).lower())

    out_files = []  # (filename, bytes)

    for hod in unique_hods:
        sub_df = df[df[hod_col] == hod].copy()
        if sub_df.empty:
            continue

        wb = load_workbook(io.BytesIO(template_file.getvalue()), data_only=False, keep_vba=keep_vba)
        data_ws, header_row, subheader_row = detect_data_sheet_and_header(wb)

        # enforce unified names on HOD sheet
        if perhod_rename_map:
            rename_headers_in_sheet(data_ws, header_row, perhod_rename_map)

        if normalize_headers:
            normalize_header_styles(data_ws, header_row, subheader_row=subheader_row)

        tmpl_pos = header_positions(data_ws, header_row)
        model_row = header_row + 1
        model_formulas = collect_model_formulas(data_ws, model_row)
        model_numfmts  = collect_model_number_formats(data_ws, model_row)
        target_headers = list(tmpl_pos.keys())

        clear_cells_below(data_ws, header_row, tmpl_pos, max_rows=50000)
        master_rows = df_to_rows_via_mapping(sub_df, target_headers, mapping_perhod, numeric_targets=numeric_targets)
        write_rows(data_ws, tmpl_pos, master_rows, model_formulas, model_row, model_numfmts, initialize_model_row=True)

        if master_rows:
            min_col = min(c for (c, _) in tmpl_pos.values())
            max_col = max(c for (c, _) in tmpl_pos.values())
            last_row = model_row + len(master_rows) - 1
            expand_table_if_present(data_ws, header_row, last_row, min_col=min_col, max_col=max_col)

        # HR_View from formulas
        hr_cols_from_fx = extract_hrview_columns_from_formulas(model_formulas)
        build_or_refresh_hr_view_table(wb, hr_cols_from_fx, master_rows, numeric_targets=numeric_targets)

        # L+2 sheets
        unique_l2 = sorted([x for x in sub_df[l2_col].dropna().unique() if str(x).strip() != ""], key=lambda s: str(s).lower())
        for mname in unique_l2:
            l2_df = sub_df[sub_df[l2_col] == mname].copy()
            if l2_df.empty:
                continue

            sheet_name = sanitize_sheet_name(f"L2 - {mname}")
            ws = wb.create_sheet(title=sheet_name)

            clone_column_widths(data_ws, ws)
            if subheader_row:
                copy_row_values_and_styles(data_ws, ws, subheader_row, subheader_row)
            copy_row_values_and_styles(data_ws, ws, header_row, header_row)
            if normalize_headers:
                normalize_header_styles(ws, header_row, subheader_row=subheader_row)

            tmpl_pos_new = header_positions(ws, header_row)
            model_row_new = header_row + 1

            for col_idx, f in model_formulas.items():
                ws.cell(model_row_new, col_idx).value = f
                try:
                    ws.cell(model_row_new, col_idx).number_format = model_numfmts.get(col_idx, ws.cell(model_row_new, col_idx).number_format)
                except Exception:
                    pass

            rows = df_to_rows_via_mapping(l2_df, target_headers, mapping_perhod, numeric_targets=numeric_targets)
            write_rows(ws, tmpl_pos_new, rows, model_formulas, model_row_new, model_numfmts, initialize_model_row=False)

        safe_hod = str(hod).strip() or "Unknown_HOD"
        out_name = f"{output_name_pattern.replace('{HOD}', safe_hod)}{out_ext}"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        out_files.append((out_name, buf.read()))

    return out_files

def build_master_file(
    df,
    master_template_file,
    mapping_master,
    numeric_targets,
    all_mgrs_columns=None,          # if None, we use the fixed list
    all_mgrs_renames=None,          # not used now; kept for API compatibility
    master_rename_map=None,
    normalize_headers=True,
    master_filename="MasterFile",
):
    """
    Build MasterFile with:
      - Master data sheet (all template columns)
      - All_Managers sheet with the fixed column list (subset of Master)
    """
    ctx = load_template_context(master_template_file)
    wb = ctx["wb"]
    data_ws = ctx["data_ws"]
    header_row = ctx["header_row"]
    subheader_row = ctx["subheader_row"]
    tmpl_pos = ctx["tmpl_pos"]
    model_row = ctx["model_row"]
    model_formulas = ctx["model_formulas"]
    model_numfmts = ctx["model_numfmts"]
    keep_vba = ctx["keep_vba"]
    out_ext = ctx["ext"]

    # enforce unified names on master data sheet
    if master_rename_map:
        rename_headers_in_sheet(data_ws, header_row, master_rename_map)
        tmpl_pos = header_positions(data_ws, header_row)  # refresh after rename

    if normalize_headers:
        normalize_header_styles(data_ws, header_row, subheader_row=subheader_row)

    target_headers = list(tmpl_pos.keys())

    # --- Master data sheet
    clear_cells_below(data_ws, header_row, tmpl_pos, max_rows=100000)
    all_rows = df_to_rows_via_mapping(df, target_headers, mapping_master, numeric_targets=numeric_targets)
    write_rows(data_ws, tmpl_pos, all_rows, model_formulas, model_row, model_numfmts, initialize_model_row=True)

    if all_rows:
        min_col = min(c for (c, _) in tmpl_pos.values())
        max_col = max(c for (c, _) in tmpl_pos.values())
        last_row = model_row + len(all_rows) - 1
        expand_table_if_present(data_ws, header_row, last_row, min_col=min_col, max_col=max_col)

    # --- All_Managers (fixed subset, based on Master rows)
    if all_mgrs_columns is None:
        all_mgrs_columns = ALL_MANAGERS_FIXED_COLUMNS
    rename_map = {} if all_mgrs_renames is None else all_mgrs_renames
    final_headers = [rename_map.get(h, h) for h in all_mgrs_columns]

    # build sheet
    if ALL_MGRS_SHEET in wb.sheetnames:
        ws = wb[ALL_MGRS_SHEET]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=ALL_MGRS_SHEET)
    remove_table_if_exists(ws, ALL_MGRS_TABLE)

    # headers
    for j, colname in enumerate(final_headers, start=1):
        ws.cell(1, j).value = colname

    # data: take from Master rows (all_rows), use original (unrenamed) keys from target_headers
    for i, row in enumerate(all_rows, start=2):
        for j, orig_header in enumerate(all_mgrs_columns, start=1):
            val = row.get(orig_header, "")
            if isinstance(val, str) and orig_header in (numeric_targets or []):
                val = coerce_excel_number(val)
            ws.cell(i, j).value = val

    # table
    last_col_letter = get_column_letter(len(final_headers))
    last_row = 1 + len(all_rows)
    table_ref = f"A1:{last_col_letter}{max(last_row, 1)}"
    table = Table(displayName=ALL_MGRS_TABLE, ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    out_name = f"{master_filename}{out_ext}"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return out_name, buf.read()
